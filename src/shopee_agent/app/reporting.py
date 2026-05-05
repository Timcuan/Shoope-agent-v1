from __future__ import annotations

import hashlib
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from shopee_agent.contracts.reporting import AuditTransaction, ReportRequest, ReportResult
from shopee_agent.persistence.repositories import ExportLogData, ExportRepository

# --- Styling constants matching auditshopeedef.xlsx ---
HEADER_FONT = Font(name="Calibri", bold=True, size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT_WHITE = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
THIN_BORDER_SIDE = Side(style="thin")
THIN_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")
IDR_FORMAT = '#,##0'

MONTH_NAMES = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April", 5: "Mei", 6: "Juni",
    7: "Juli", 8: "Agustus", 9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}


class ReportingAgent:
    def __init__(self, export_repo: ExportRepository, output_dir: Path | str = "./data/exports") -> None:
        self.export_repo = export_repo
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_audit_workbook(self, request: ReportRequest, activity_logs: list = None) -> ReportResult:
        """
        Generate a comprehensive audit workbook.
        Includes summary, daily breakdown, monthly sheets, and detailed activity logs.
        """
        wb = Workbook()
        
        # 1. Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "RINGKASAN"
        self._build_summary_sheet(ws_summary, request)
        
        # 2. Daily Breakdown Sheet (NEW)
        ws_daily = wb.create_sheet("BREAKDOWN HARIAN")
        self._build_daily_breakdown_sheet(ws_daily, request)
        
        # 3. Monthly Sheets (12 Months per requirement)
        for m in range(1, 13):
            sheet_name = MONTH_NAMES[m]
            ws = wb.create_sheet(sheet_name)
            self._build_month_sheet(ws, m, request)
        
        # 4. Detailed Activity Ledger (NEW - if logs provided)
        if activity_logs:
            ws_logs = wb.create_sheet("LOG AKTIVITAS")
            self._populate_activity_sheet(ws_logs, activity_logs, request.creator)

        # Write the file
        month_name = MONTH_NAMES.get(request.month, "unknown")
        filename = f"audit_{request.shop_id}_{request.year}_{month_name}.xlsx"
        file_path = self.output_dir / filename
        wb.save(str(file_path))

        # Compute checksum
        checksum = hashlib.sha256(file_path.read_bytes()).hexdigest()
        export_id = f"exp_{uuid.uuid4().hex}"

        period_start = datetime(request.year, request.month, 1)
        # Last day of requested month
        if request.month == 12:
            period_end = datetime(request.year + 1, 1, 1)
        else:
            period_end = datetime(request.year, request.month + 1, 1)

        # Log to database
        self.export_repo.log_export(ExportLogData(
            export_id=export_id,
            report_type="audit_workbook",
            shop_id=request.shop_id,
            period_start=period_start,
            period_end=period_end,
            file_path=str(file_path),
            checksum=checksum,
            creator=request.creator,
        ))

        # Calculate Totals
        total_rev = sum(t.order_amount for t in request.transactions)
        total_rec = sum(t.dana_diterima for t in request.transactions)

        return ReportResult(
            export_id=export_id,
            file_path=str(file_path),
            checksum=checksum,
            report_type="audit_workbook",
            period_start=period_start,
            period_end=period_end,
            row_count=len(request.transactions),
            total_revenue=total_rev,
            total_received=total_rec,
        )

    def generate_restock_workbook(self, shop_id: str, proposals: list[dict], creator: str) -> ReportResult:
        """Generate a restock/PO Excel file for suppliers."""
        wb = Workbook()
        ws = wb.active
        ws.title = "RESTOCK PLAN"
        
        headers = ["ITEM ID", "SKU", "NAMA PRODUK", "STOK SAAT INI", "VELOCITY/HARI", "JUMLAH RESTOCK", "PRIORITAS"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            
        for p in proposals:
            ws.append([
                p["item_id"], p["sku"], p["name"], p["current_stock"],
                p["velocity"], p["restock_qty"], p["priority"]
            ])
            
        filename = f"restock_{shop_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        file_path = self.output_dir / filename
        wb.save(str(file_path))
        
        checksum = hashlib.sha256(file_path.read_bytes()).hexdigest()
        export_id = f"restock_{uuid.uuid4().hex[:8]}"
        
        # Log export
        self.export_repo.log_export(ExportLogData(
            export_id=export_id,
            report_type="restock_plan",
            shop_id=shop_id,
            period_start=datetime.now(),
            period_end=datetime.now(),
            file_path=str(file_path),
            checksum=checksum,
            creator=creator,
        ))
        
        return ReportResult(
            export_id=export_id,
            file_path=str(file_path),
            checksum=checksum,
            report_type="restock_plan",
            period_start=datetime.now(),
            period_end=datetime.now(),
            row_count=len(proposals),
        )

    def generate_daily_ledger(self, shop_id: str, logs: list, creator: str) -> ReportResult:
        """
        Generates a hyper-detailed activity ledger.
        Groups logs by day, creating a separate sheet for each day.
        """
        wb = Workbook()
        del wb["Sheet"] # Remove default sheet
        
        # Group logs by date
        from collections import defaultdict
        daily_logs = defaultdict(list)
        for log in logs:
            day_str = log.created_at.strftime('%Y-%m-%d')
            daily_logs[day_str].append(log)
            
        # Sort dates
        sorted_dates = sorted(daily_logs.keys())
        
        for day in sorted_dates:
            ws = wb.create_sheet(title=day)
            
            # Header
            ws.merge_cells("A1:E1")
            ws["A1"] = f"DAILY AUDIT LEDGER - {day}"
            ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
            ws["A1"].fill = HEADER_FILL
            ws["A1"].alignment = CENTER
            
            headers = ["TIME", "TYPE", "ACTIVITY", "SEVERITY", "OPERATOR"]
            ws.append([]) # Spacer
            ws.append(headers)
            header_row = 3
            for cell in ws[header_row]:
                cell.font = HEADER_FONT_WHITE
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER
                
            # Content
            for i, log in enumerate(daily_logs[day], start=1):
                row = [
                    log.created_at.strftime('%H:%M:%S'),
                    log.activity_type.upper(),
                    log.message,
                    log.severity.upper(),
                    creator
                ]
                ws.append(row)
                
                # Dynamic cell styling
                curr_row = header_row + i
                for cell in ws[curr_row]:
                    cell.border = THIN_BORDER
                    if log.severity == "error":
                        cell.fill = RED_FILL
                        cell.font = RED_FONT
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

        filename = f"LEDGER_{shop_id}_{datetime.now().strftime('%Y%m')}.xlsx"
        file_path = self.output_dir / filename
        wb.save(str(file_path))
        
        checksum = hashlib.sha256(file_path.read_bytes()).hexdigest()
        export_id = f"ledger_{uuid.uuid4().hex[:8]}"
        
        # Log export
        self.export_repo.log_export(ExportLogData(
            export_id=export_id,
            report_type="daily_ledger",
            shop_id=shop_id,
            period_start=datetime.now(),
            period_end=datetime.now(),
            file_path=str(file_path),
            checksum=checksum,
            creator=creator,
        ))
        
        return ReportResult(
            export_id=export_id,
            file_path=str(file_path),
            checksum=checksum,
            report_type="daily_ledger",
            period_start=datetime.now(),
            period_end=datetime.now(),
            row_count=len(logs)
        )

    def _build_month_sheet(self, ws, month: int, request: ReportRequest) -> None:
        # --- Row 1: Section header merges ---
        # A1:A2 → NO (will merge vertically below)
        ws.merge_cells("A1:A2")
        ws["A1"] = "NO"
        ws["A1"].font = HEADER_FONT_WHITE
        ws["A1"].fill = HEADER_FILL
        ws["A1"].alignment = CENTER

        # B1:D1 → STATUS PESANAN
        ws.merge_cells("B1:D1")
        ws["B1"] = "STATUS PESANAN"
        ws["B1"].font = HEADER_FONT_WHITE
        ws["B1"].fill = HEADER_FILL
        ws["B1"].alignment = CENTER

        # E1:F1 → DETAIL PESANAN
        ws.merge_cells("E1:F1")
        ws["E1"] = "DETAIL PESANAN"
        ws["E1"].font = HEADER_FONT_WHITE
        ws["E1"].fill = HEADER_FILL
        ws["E1"].alignment = CENTER

        # G1:I1 → BIAYA
        ws.merge_cells("G1:I1")
        ws["G1"] = "BIAYA"
        ws["G1"].font = HEADER_FONT_WHITE
        ws["G1"].fill = HEADER_FILL
        ws["G1"].alignment = CENTER

        # J1:L1 → LAIN-LAIN
        ws.merge_cells("J1:L1")
        ws["J1"] = "LAIN-LAIN"
        ws["J1"].font = HEADER_FONT_WHITE
        ws["J1"].fill = HEADER_FILL
        ws["J1"].alignment = CENTER

        # --- Row 2: Field headers ---
        field_headers = {
            "B2": "TGL TERIMA",
            "C2": "TGL KIRIM",
            "D2": "TGL SELESAI",
            "E2": "NO. PESANAN",
            "F2": "NAMA PRODUK",
            "G2": "HARGA JUAL",
            "H2": "BIAYA KOMISI",
            "I2": "BIAYA LAYANAN",
            "J2": "TOTAL POTONGAN",
            "K2": "EST. PENDAPATAN",
            "L2": "DANA DITERIMA",
            "M2": "SELISIH",
            "N2": "STATUS"
        }
        for cell_ref, label in field_headers.items():
            ws[cell_ref] = label
            ws[cell_ref].font = HEADER_FONT_WHITE
            ws[cell_ref].fill = HEADER_FILL
            ws[cell_ref].alignment = CENTER
            ws[cell_ref].border = THIN_BORDER

        # --- Rows 3:202: Transaction data (only fill current month's rows) ---
        DATA_START = 3
        DATA_END = 202  # 200 rows per month per spec

        # Filter transactions for this month
        txns_for_month: list[AuditTransaction] = [
            t for t in request.transactions
            if t.received_at and t.received_at.month == month
        ] if month == request.month else []

        admin_rate = request.admin_rate

        for i in range(DATA_START, DATA_END + 1):
            row_idx = i - DATA_START  # 0-indexed into txns_for_month
            ws[f"A{i}"] = i - DATA_START + 1  # sequential row number

            if row_idx < len(txns_for_month):
                t = txns_for_month[row_idx]
                ws[f"B{i}"] = t.received_at
                ws[f"C{i}"] = t.shipped_at
                ws[f"D{i}"] = t.completed_at
                ws[f"E{i}"] = t.order_sn
                ws[f"F{i}"] = t.order_label
                ws[f"G{i}"] = t.order_amount
                
                # Biaya Marketplace (Commission)
                ws[f"H{i}"] = t.biaya_admin if t.biaya_admin > 0 else f"=G{i}*{admin_rate}"
                # Biaya Layanan (Xtra etc)
                ws[f"I{i}"] = t.biaya_layanan
                
                # Formulas
                ws[f"J{i}"] = f"=H{i}+I{i}"
                ws[f"K{i}"] = f"=G{i}-J{i}"
                ws[f"L{i}"] = t.dana_diterima
                ws[f"M{i}"] = f"=L{i}-K{i}"
                ws[f"N{i}"] = f"=IF(ABS(M{i})<100, \"MATCH\", \"CHECK\")"
                
                # Styles
                for col in ["G", "H", "I", "J", "K", "L", "M"]:
                    ws[f"{col}{i}"].number_format = IDR_FORMAT
                
                if t.dana_diterima < (t.order_amount * 0.8): # Huge fee warning
                    ws[f"M{i}"].fill = RED_FILL
                    ws[f"M{i}"].font = RED_FONT
            else:
                # Fill empty rows with formulas for user manual input
                ws[f"H{i}"] = f"=IF(G{i}<>\"\", G{i}*{admin_rate}, \"\")"
                ws[f"I{i}"] = f"=IF(G{i}<>\"\", G{i}-H{i}, \"\")"
                ws[f"K{i}"] = f"=IF(J{i}<>\"\", J{i}-I{i}, \"\")"
                ws[f"L{i}"] = f"=IF(G{i}>0, J{i}/G{i}, \"\")"

        # --- Row 203: Total row ---
        TOTAL_ROW = 203
        ws[f"A{TOTAL_ROW}"] = "TOTAL"
        ws[f"A{TOTAL_ROW}"].font = HEADER_FONT
        for col in ["G", "H", "I", "J", "K", "L", "M"]:
            ws[f"{col}{TOTAL_ROW}"] = f"=SUM({col}{DATA_START}:{col}{DATA_END})"
            ws[f"{col}{TOTAL_ROW}"].font = HEADER_FONT
            ws[f"{col}{TOTAL_ROW}"].number_format = IDR_FORMAT
            ws[f"{col}{TOTAL_ROW}"].border = THIN_BORDER
        
        # Column widths
        col_widths = {"A": 6, "B": 12, "C": 12, "D": 12, "E": 14, "F": 25,
                      "G": 14, "H": 12, "I": 12, "J": 14, "K": 16, "L": 16, "M": 12, "N": 12}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

    def format_telegram_summary(self, result: ReportResult) -> str:
        """Generate a premium Markdown summary for Telegram."""
        selisih = result.total_revenue - result.total_received
        selisih_pct = (selisih / result.total_revenue * 100) if result.total_revenue > 0 else 0
        
        status_icon = "🟢" if selisih_pct < 5 else "🟡" if selisih_pct < 10 else "🔴"
        
        return (
            f"📑 **REKAP AUDIT TERBIT**\n"
            f"━━━━━━━━━━━━━━━\n\n"
            f"🏪 **Toko:** `{result.shop_id}`\n"
            f"📅 **Periode:** {result.period_start.strftime('%B %Y')}\n\n"
            f"💰 **Total Omzet:** `Rp {result.total_revenue:,.0f}`\n"
            f"📥 **Dana Diterima:** `Rp {result.total_received:,.0f}`\n"
            f"💸 **Total Potongan:** `Rp {selisih:,.0f}` (`{selisih_pct:.1f}%`)\n\n"
            f"📦 **Total Transaksi:** `{result.row_count}`\n"
            f"📊 **Health Status:** {status_icon} *HEALTHY*" if selisih_pct < 5 else f"📊 **Health Status:** {status_icon} *CHECK FEES*" 
            f"\n━━━━━━━━━━━━━━━\n"
            f"🔗 ID: `{result.export_id}`\n"
            f"🔐 Checksum: `{result.checksum[:8]}...`"
        )

    def _build_summary_sheet(self, ws, request: ReportRequest) -> None:
        """Build a high-level KPI sheet at the start of the workbook."""
        ws["A1"] = "RINGKASAN AUDIT TOKO"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")
        
        ws["A3"] = "Toko ID"
        ws["B3"] = request.shop_id
        ws["A4"] = "Tahun"
        ws["B4"] = request.year
        ws["A5"] = "Bulan"
        ws["B5"] = MONTH_NAMES[request.month]
        
        for i in range(3, 6):
            ws[f"A{i}"].font = HEADER_FONT
            
        # KPI Box
        headers = ["METRIK", "NILAI"]
        for idx, h in enumerate(headers):
            cell = ws.cell(row=7, column=idx+1, value=h)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            
        total_rev = sum(t.order_amount for t in request.transactions)
        total_rec = sum(t.dana_diterima for t in request.transactions)
        
        data = [
            ("Total Omzet (GMV)", total_rev),
            ("Total Dana Diterima (Nett)", total_rec),
            ("Total Potongan (Fees/Admin)", total_rev - total_rec),
            ("Jumlah Transaksi", len(request.transactions)),
        ]
        
        for idx, (label, val) in enumerate(data):
            ws.cell(row=8+idx, column=1, value=label)
            cell_val = ws.cell(row=8+idx, column=2, value=val)
            if isinstance(val, (int, float)):
                cell_val.number_format = IDR_FORMAT
        
        # Add a small note
        ws["A14"] = "Laporan ini dihasilkan secara otomatis oleh Shopee Orchestrator v1.0"
        ws["A14"].font = Font(italic=True, size=9)
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _build_daily_breakdown_sheet(self, ws, request: ReportRequest) -> None:
        """Creates a day-by-day revenue and income summary."""
        headers = ["TANGGAL", "JUMLAH ORDER", "TOTAL PENJUALAN", "DANA DITERIMA"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            
        # Group by date
        from collections import defaultdict
        daily_stats = defaultdict(lambda: {"count": 0, "rev": 0, "rec": 0})
        for t in request.transactions:
            if t.received_at:
                day = t.received_at.strftime("%Y-%m-%d")
                daily_stats[day]["count"] += 1
                daily_stats[day]["rev"] += t.order_amount
                daily_stats[day]["rec"] += t.dana_diterima
                
        for day in sorted(daily_stats.keys()):
            stats = daily_stats[day]
            ws.append([day, stats["count"], stats["rev"], stats["rec"]])
            
        # Add Totals Row
        row_idx = ws.max_row + 1
        ws[f"A{row_idx}"] = "TOTAL"
        ws[f"A{row_idx}"].font = Font(bold=True)
        ws[f"B{row_idx}"] = f"=SUM(B2:B{row_idx-1})"
        ws[f"C{row_idx}"] = f"=SUM(C2:C{row_idx-1})"
        ws[f"D{row_idx}"] = f"=SUM(D2:D{row_idx-1})"
        
        for col in ["C", "D"]:
            for i in range(2, row_idx + 1):
                ws[f"{col}{i}"].number_format = IDR_FORMAT
        
        self._auto_size_columns(ws)

    def _populate_activity_sheet(self, ws, logs: list, creator: str) -> None:
        """Populates detailed activity logs with color coding."""
        headers = ["TIMESTAMP", "TYPE", "MESSAGE", "SEVERITY"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            
        for log in logs:
            ws.append([
                log.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                log.activity_type.upper(),
                log.message,
                log.severity.upper()
            ])
            
            # Highlight errors
            if log.severity == "error":
                for cell in ws[ws.max_row]:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
                    
        self._auto_size_columns(ws)

    def _auto_size_columns(self, ws) -> None:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 2, 60)
