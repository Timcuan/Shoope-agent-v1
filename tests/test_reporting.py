import hashlib
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

from shopee_agent.app.reporting import ReportingAgent
from shopee_agent.contracts.reporting import AuditTransaction, ReportRequest
from shopee_agent.persistence.repositories import ExportRepository


# ---- Helpers ----

class FakeExportRepo:
    """In-memory stand-in for ExportRepository."""
    def __init__(self):
        self.logged = []

    def log_export(self, data) -> None:
        self.logged.append(data)

    def get_recent_exports(self, report_type, limit=10):
        return self.logged


def make_agent(tmp_path: Path) -> tuple[ReportingAgent, FakeExportRepo]:
    repo = FakeExportRepo()
    agent = ReportingAgent(export_repo=repo, output_dir=tmp_path)
    return agent, repo


# ---- Tests ----

def test_generate_audit_workbook_creates_file(tmp_path: Path) -> None:
    agent, repo = make_agent(tmp_path)
    req = ReportRequest(shop_id="shop1", year=2026, month=5)

    result = agent.generate_audit_workbook(req)

    assert Path(result.file_path).exists(), "Excel file must be written to disk"
    assert result.report_type == "audit_workbook"
    assert result.row_count == 0  # no transactions provided


def test_audit_workbook_has_comprehensive_sheets(tmp_path: Path) -> None:
    agent, _ = make_agent(tmp_path)
    req = ReportRequest(shop_id="shop1", year=2026, month=5)
    result = agent.generate_audit_workbook(req)

    wb = openpyxl.load_workbook(result.file_path)
    # RINGKASAN, BREAKDOWN HARIAN, 12 Months = 14 sheets
    # LOG AKTIVITAS is only added if activity_logs provided, so here it's 14.
    assert len(wb.sheetnames) == 14
    assert wb.sheetnames[0] == "RINGKASAN"
    assert "BREAKDOWN HARIAN" in wb.sheetnames
    assert "Mei" in wb.sheetnames


def test_audit_workbook_section_headers(tmp_path: Path) -> None:
    agent, _ = make_agent(tmp_path)
    req = ReportRequest(shop_id="shop1", year=2026, month=5)
    result = agent.generate_audit_workbook(req)

    wb = openpyxl.load_workbook(result.file_path)
    may_sheet = wb["Mei"]

    assert may_sheet["A1"].value == "NO"
    assert may_sheet["B1"].value == "STATUS PESANAN"
    assert may_sheet["E1"].value == "DETAIL PESANAN"
    assert may_sheet["G1"].value == "BIAYA"
    assert may_sheet["J1"].value == "LAIN-LAIN"


def test_audit_workbook_field_headers(tmp_path: Path) -> None:
    agent, _ = make_agent(tmp_path)
    req = ReportRequest(shop_id="shop1", year=2026, month=5)
    result = agent.generate_audit_workbook(req)

    wb = openpyxl.load_workbook(result.file_path)
    ws = wb["Mei"]

    expected = {
        "B2": "TGL TERIMA", "C2": "TGL KIRIM", "D2": "TGL SELESAI",
        "E2": "NO. PESANAN", "F2": "NAMA PRODUK",
        "G2": "HARGA JUAL", "H2": "BIAYA KOMISI", "I2": "BIAYA LAYANAN",
        "J2": "TOTAL POTONGAN", "K2": "EST. PENDAPATAN",
        "L2": "DANA DITERIMA", "M2": "SELISIH", "N2": "STATUS"
    }
    for cell_ref, expected_label in expected.items():
        assert ws[cell_ref].value == expected_label, f"{cell_ref} should be {expected_label!r}"


def test_audit_workbook_total_row_formulas(tmp_path: Path) -> None:
    agent, _ = make_agent(tmp_path)
    req = ReportRequest(shop_id="shop1", year=2026, month=5)
    result = agent.generate_audit_workbook(req)

    wb = openpyxl.load_workbook(result.file_path)
    ws = wb["Mei"]

    assert "SUM" in str(ws["G203"].value)
    assert "SUM" in str(ws["L203"].value)
    assert "SUM" in str(ws["M203"].value)


def test_audit_workbook_with_transactions(tmp_path: Path) -> None:
    agent, repo = make_agent(tmp_path)
    txns = [
        AuditTransaction(
            row_no=1,
            received_at=date(2026, 5, 1),
            shipped_at=date(2026, 5, 2),
            completed_at=date(2026, 5, 5),
            order_label="SPX",
            order_sn="250501ABC",
            order_amount=100_000.0,
            admin_rate=0.02,
            dana_diterima=98_000.0,
            keterangan="normal",
        )
    ]
    req = ReportRequest(shop_id="shop1", year=2026, month=5, transactions=txns)
    result = agent.generate_audit_workbook(req)

    assert result.row_count == 1

    wb = openpyxl.load_workbook(result.file_path)
    ws = wb["Mei"]
    assert ws["E3"].value == "250501ABC"
    assert ws["G3"].value == 100_000.0
    assert ws["L3"].value == 98_000.0


def test_audit_workbook_export_logged(tmp_path: Path) -> None:
    agent, repo = make_agent(tmp_path)
    req = ReportRequest(shop_id="shop1", year=2026, month=5, creator="test_operator")
    result = agent.generate_audit_workbook(req)

    assert len(repo.logged) == 1
    log = repo.logged[0]
    assert log.report_type == "audit_workbook"
    assert log.shop_id == "shop1"
    assert log.creator == "test_operator"
    assert log.checksum == result.checksum


def test_audit_workbook_checksum_stable(tmp_path: Path) -> None:
    """Same request must produce same checksum."""
    agent, _ = make_agent(tmp_path)
    req = ReportRequest(shop_id="shop1", year=2026, month=5)
    r1 = agent.generate_audit_workbook(req)
    checksum1 = hashlib.sha256(Path(r1.file_path).read_bytes()).hexdigest()
    assert r1.checksum == checksum1
